import os
import time
from datetime import date, datetime, timedelta
from pathlib import Path
import numpy as np
import pandas as pd
from sqlalchemy import select, func, and_
from sqlalchemy.orm import Session
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import Security, RawPrice, AdjustedPrice, MarketCap


def get_closest_trading_date(session: Session, target_date: date) -> date:
    """
    Find the closest trading date in the database on or before target_date.
    """
    max_date = session.execute(
        select(func.max(RawPrice.trade_date))
        .where(RawPrice.trade_date <= target_date)
    ).scalar()
    
    if not max_date:
        # If no dates before target_date, fallback to the oldest available date
        max_date = session.execute(select(func.min(RawPrice.trade_date))).scalar()
        
    return max_date


def get_next_available_trading_date(all_dates: list[date], target_date: date) -> date:
    """
    Find the first date in the sorted all_dates list that is >= target_date.
    If none is found, return the last date in all_dates (the latest date).
    """
    for d in all_dates:
        if d >= target_date:
            return d
    return all_dates[-1] if all_dates else target_date


def run_sharpe_screener(
    session: Session,
    target_date: date,
    long_months: int = 6,
    short_months: int = 3,
    mcap_filter_cr: float = 1000.0,
    roc_annual_filter: float = 6.5,
    turnover_filter_cr: float = 1.0
) -> pd.DataFrame:
    """
    Runs the Sharpe Based Screener:
    1. Filters active stocks with Market Cap >= mcap_filter_cr on target_date.
    2. Calculates Annual ROC and Median Daily Turnover. Filters out stocks failing base limits.
    3. Calculates Sharpe ratios (Long/Short), DMAs, 52W High Proximity, and Circuit Hits on passed subset.
    4. Ranks stocks by combined Sharpe rank sum.
    """
    t_start = time.time()
    
    # Ensure target_date is a valid trading date (or closest prior)
    t_date = get_closest_trading_date(session, target_date)
    if not t_date:
        logger.warning("No trading dates found in the database.")
        return pd.DataFrame()
        
    logger.info(f"Running Sharpe Screener relative to trading date: {t_date}")
    
    # Step 1: Filter active stocks with Market Cap >= mcap_filter_cr
    mcap_threshold_bps = mcap_filter_cr * 10_000_000.0  # Convert Cr to absolute Rupees
    
    mcap_query = (
        select(
            Security.id.label("security_id"),
            Security.symbol,
            Security.company_name,
            Security.isin,
            Security.industry,
            MarketCap.market_cap
        )
        .join(MarketCap, Security.id == MarketCap.security_id)
        .where(Security.security_type == "STOCK")
        .where(Security.is_active == True)
        .where(Security.is_delisted == False)
        .where(MarketCap.trade_date == t_date)
        .where(MarketCap.market_cap >= mcap_threshold_bps)
    )
    
    mcap_rows = session.execute(mcap_query).all()
    if not mcap_rows:
        logger.warning(f"No stocks passed the Market Cap filter (>= {mcap_filter_cr} Cr) on {t_date}.")
        return pd.DataFrame()
        
    passing_stocks = {
        r.security_id: {
            "symbol": r.symbol,
            "company_name": r.company_name or "-",
            "isin": r.isin or "-",
            "industry": r.industry or "-",
            "market_cap_cr": round(float(r.market_cap) / 10_000_000.0, 2)
        }
        for r in mcap_rows
    }
    
    logger.info(f"Market Cap filter passed: {len(passing_stocks)} stocks.")
    
    # Get all distinct trading dates up to target date T, sorted ascending
    dates_query = (
        select(RawPrice.trade_date)
        .distinct()
        .where(RawPrice.trade_date <= t_date)
        .order_by(RawPrice.trade_date.desc())
        .limit(260)  # Load last 260 trading days for indicator calc
    )
    trading_dates = sorted(list(session.execute(dates_query).scalars().all()))
    
    if len(trading_dates) < 252:
        logger.warning(f"Insufficient trading history in database. Need at least 252 days, found {len(trading_dates)}.")
        return pd.DataFrame()
        
    start_history_date = trading_dates[0]
    
    # Calculate Calendar Offset Target Dates
    t_dt = pd.to_datetime(t_date)
    target_3m_dt = (t_dt - pd.DateOffset(months=short_months)).date()
    target_6m_dt = (t_dt - pd.DateOffset(months=long_months)).date()
    target_12m_dt = (t_dt - pd.DateOffset(months=12)).date()
    
    # Resolve exact database trading dates using next available fallback
    t_3m = get_next_available_trading_date(trading_dates, target_3m_dt)
    t_6m = get_next_available_trading_date(trading_dates, target_6m_dt)
    t_12m = get_next_available_trading_date(trading_dates, target_12m_dt)
    
    logger.info(f"Target Dates Resolved:")
    logger.info(f"  3M Lookback Date  : {t_3m} (requested: {target_3m_dt})")
    logger.info(f"  6M Lookback Date  : {t_6m} (requested: {target_6m_dt})")
    logger.info(f"  12M Lookback Date : {t_12m} (requested: {target_12m_dt})")
    
    # Step 2: Bulk load raw and adjusted prices for the passing stocks
    price_query = (
        select(
            AdjustedPrice.security_id,
            AdjustedPrice.trade_date,
            AdjustedPrice.adj_close,
            RawPrice.close.label("raw_close"),
            RawPrice.prev_close.label("raw_prev_close"),
            RawPrice.volume
        )
        .join(
            RawPrice,
            (AdjustedPrice.security_id == RawPrice.security_id) & (AdjustedPrice.trade_date == RawPrice.trade_date)
        )
        .where(AdjustedPrice.security_id.in_(list(passing_stocks.keys())))
        .where(AdjustedPrice.trade_date >= start_history_date)
        .where(AdjustedPrice.trade_date <= t_date)
        .order_by(AdjustedPrice.security_id, AdjustedPrice.trade_date.asc())
    )
    
    price_rows = session.execute(price_query).all()
    if not price_rows:
        logger.warning("No price history found for the filtered stocks.")
        return pd.DataFrame()
        
    df_prices = pd.DataFrame([{
        "security_id": r.security_id,
        "trade_date": r.trade_date,
        "adj_close": float(r.adj_close),
        "raw_close": float(r.raw_close),
        "raw_prev_close": float(r.raw_prev_close) if r.raw_prev_close else None,
        "volume": int(r.volume)
    } for r in price_rows])
    
    results = []
    
    # Process each stock group
    for sec_id, grp in df_prices.groupby("security_id"):
        grp = grp.sort_values("trade_date")
        
        # Ensure target date is present in the group
        dates_in_grp = grp["trade_date"].values
        if len(dates_in_grp) == 0 or dates_in_grp[-1] != t_date:
            continue
            
        # We need sufficient history (minimum 252 trading days)
        if len(grp) < 252:
            continue
            
        adj_close_T = grp.iloc[-1]["adj_close"]
        raw_close_T = grp.iloc[-1]["raw_close"]
        
        # 1. Annual ROC (Annual Return): (adj_close_T - adj_close_12m) / adj_close_12m * 100
        idx_12m = np.where(dates_in_grp >= t_12m)[0]
        if len(idx_12m) == 0:
            continue
        adj_close_12m = grp.iloc[idx_12m[0]]["adj_close"]
        if adj_close_12m <= 0:
            continue
        roc_annual = ((adj_close_T - adj_close_12m) / adj_close_12m) * 100
        
        # Base filter check for Annual ROC
        if roc_annual < roc_annual_filter:
            continue
            
        # 2. Median Daily Turnover: median of (raw_close * volume) over last 252 days
        last_252 = grp.iloc[-252:]
        daily_turnover = last_252["raw_close"] * last_252["volume"]
        median_turnover_cr = round(float(daily_turnover.median()) / 10_000_000.0, 4)
        
        # Base filter check for Turnover
        if median_turnover_cr < turnover_filter_cr:
            continue
            
        # Stocks passing here are the official base-filtered candidates!
        # Now run Sharpe ratios and other indicators.
        
        # 3. Daily Returns on Adjusted Close for Sharpe calculations
        adj_returns = grp["adj_close"].pct_change().dropna()
        
        # Sharpe 3M (Short Window)
        idx_3m = np.where(dates_in_grp >= t_3m)[0]
        sharpe_3 = None
        roc_3 = None
        if len(idx_3m) > 0:
            start_3m_idx = idx_3m[0]
            # daily returns slice for the 3M window
            # index offset needs to match the slice of daily returns
            slice_dates = dates_in_grp[start_3m_idx:]
            if len(slice_dates) >= 2:
                ret_3m = adj_returns[adj_returns.index >= grp.index[start_3m_idx + 1]]
                std_3 = ret_3m.std(ddof=1)
                sharpe_3 = round(ret_3m.mean() / std_3, 4) if std_3 > 0 else None
                
                # 3M ROC
                adj_close_3m = grp.iloc[start_3m_idx]["adj_close"]
                if adj_close_3m > 0:
                    roc_3 = round(((adj_close_T - adj_close_3m) / adj_close_3m) * 100, 2)
                    
        # Sharpe 6M (Long Window)
        idx_6m = np.where(dates_in_grp >= t_6m)[0]
        sharpe_6 = None
        roc_6 = None
        if len(idx_6m) > 0:
            start_6m_idx = idx_6m[0]
            slice_dates = dates_in_grp[start_6m_idx:]
            if len(slice_dates) >= 2:
                ret_6m = adj_returns[adj_returns.index >= grp.index[start_6m_idx + 1]]
                std_6 = ret_6m.std(ddof=1)
                sharpe_6 = round(ret_6m.mean() / std_6, 4) if std_6 > 0 else None
                
                # 6M ROC
                adj_close_6m = grp.iloc[start_6m_idx]["adj_close"]
                if adj_close_6m > 0:
                    roc_6 = round(((adj_close_T - adj_close_6m) / adj_close_6m) * 100, 2)
                    
        # 4. DMA levels on adjusted close
        dma_20 = round(float(grp["adj_close"].iloc[-20:].mean()), 2)
        dma_50 = round(float(grp["adj_close"].iloc[-50:].mean()), 2)
        dma_100 = round(float(grp["adj_close"].iloc[-100:].mean()), 2)
        dma_200 = round(float(grp["adj_close"].iloc[-200:].mean()), 2)
        
        # 5. 52-Week High Proximity
        week_52_high = round(float(last_252["adj_close"].max()), 2)
        away_52wh = round(((adj_close_T - week_52_high) / week_52_high) * 100, 2) if week_52_high > 0 else None
        
        # 6. Circuit Hits Detector over the last 63 trading days (~3 months)
        last_63 = grp.iloc[-63:]
        circuit_hits = 0
        bands = [5.0, 10.0, 20.0]
        tol = 0.025  # 0.025% tolerance
        
        for _, r_row in last_63.iterrows():
            prev_c = r_row["raw_prev_close"]
            curr_c = r_row["raw_close"]
            if prev_c and prev_c > 0:
                ret_pct = ((curr_c - prev_c) / prev_c) * 100.0
                abs_ret = abs(ret_pct)
                for band in bands:
                    if (band - tol) <= abs_ret <= (band + tol):
                        circuit_hits += 1
                        break
                        
        stock_meta = passing_stocks[sec_id]
        results.append({
            "symbol": stock_meta["symbol"],
            "company_name": stock_meta["company_name"],
            "isin": stock_meta["isin"],
            "industry": stock_meta["industry"],
            "close": raw_close_T,  # raw close price
            "dma_20": dma_20,
            "dma_50": dma_50,
            "dma_100": dma_100,
            "dma_200": dma_200,
            "away_52wh": away_52wh,
            "total_circuit_hits_3m": circuit_hits,
            "sharpe_6": sharpe_6,
            "sharpe_3": sharpe_3,
            "ROC_6": roc_6,
            "ROC_3": roc_3,
            "week_52_high": week_52_high,
            "market_cap_cr": stock_meta["market_cap_cr"],
            "ROC_annual": round(roc_annual, 2),
            "median_turnover_cr": median_turnover_cr
        })
        
    df_results = pd.DataFrame(results)
    if df_results.empty:
        logger.warning("No stocks passed all calculations and filters.")
        return pd.DataFrame()
        
    # Rank stocks on Sharpe Ratios
    # Drop rows without Sharpe metrics just in case
    df_results = df_results.dropna(subset=["sharpe_6", "sharpe_3"]).copy()
    if df_results.empty:
        return pd.DataFrame()
        
    df_results["sharpe_6_rank"] = df_results["sharpe_6"].rank(ascending=False, method="first").astype(int)
    df_results["sharpe_3_rank"] = df_results["sharpe_3"].rank(ascending=False, method="first").astype(int)
    df_results["Avg_sharpe_6_3_Rank"] = df_results["sharpe_6_rank"] + df_results["sharpe_3_rank"]
    
    # Sort by combined rank sum ascending (lowest sum = best rank)
    df_results = df_results.sort_values("Avg_sharpe_6_3_Rank").reset_index(drop=True)
    
    # Add rank columns back to the columns list and organize output
    df_results.attrs["long_months"] = long_months
    df_results.attrs["short_months"] = short_months
    
    output_cols = [
        "symbol", "company_name", "industry", "close",
        "dma_20", "dma_50", "dma_100", "dma_200", "away_52wh",
        "total_circuit_hits_3m", "Avg_sharpe_6_3_Rank",
        "sharpe_6", "sharpe_3", "ROC_6", "ROC_3", "week_52_high",
        "market_cap_cr", "ROC_annual", "median_turnover_cr",
        "sharpe_6_rank", "sharpe_3_rank", "isin"
    ]
    df_results = df_results[[c for c in output_cols if c in df_results.columns]]
    
    logger.info(f"Sharpe Screener execution completed. Passed: {len(df_results)} stocks in {time.time() - t_start:.2f}s.")
    return df_results


# Styling Constants for Excel Export
HEADER_FILL = PatternFill("solid", start_color="1F4E79")   # Dark Blue
ALT_FILL = PatternFill("solid", start_color="D6E4F0")      # Soft Light Blue
FILTER_FILL = PatternFill("solid", start_color="1A5C38")    # Dark Green
ALT_FILL_FILTER = PatternFill("solid", start_color="D4EDDA") # Soft Light Green
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=9)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN_SIDE = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

COL_WIDTHS = {
    "symbol": 12,
    "company_name": 28,
    "industry": 20,
    "close": 12,
    "dma_20": 10,
    "dma_50": 10,
    "dma_100": 10,
    "dma_200": 10,
    "away_52wh": 16,
    "total_circuit_hits_3m": 18,
    "Avg_sharpe_6_3_Rank": 16,
    "sharpe_6": 12,
    "sharpe_3": 12,
    "ROC_6": 11,
    "ROC_3": 11,
    "week_52_high": 14,
    "market_cap_cr": 14,
    "ROC_annual": 13,
    "median_turnover_cr": 18,
    "sharpe_6_rank": 14,
    "sharpe_3_rank": 14,
    "isin": 16,
}

FRIENDLY_HEADERS = {
    "symbol": "Symbol",
    "company_name": "Company Name",
    "industry": "Industry",
    "close": "Close (Rs.)",
    "dma_20": "20 DMA",
    "dma_50": "50 DMA",
    "dma_100": "100 DMA",
    "dma_200": "200 DMA",
    "away_52wh": "Away from 52WH %",
    "total_circuit_hits_3m": "Circuit Hits (3M)",
    "Avg_sharpe_6_3_Rank": "Avg Sharpe Rank",
    "sharpe_6": "Sharpe 6M",
    "sharpe_3": "Sharpe 3M",
    "ROC_6": "6M ROC %",
    "ROC_3": "3M ROC %",
    "week_52_high": "52W High (Rs.)",
    "market_cap_cr": "MCAP (Cr)",
    "ROC_annual": "Annual ROC %",
    "median_turnover_cr": "Med. Turnover (Cr)",
    "sharpe_6_rank": "Sharpe 6M Rank",
    "sharpe_3_rank": "Sharpe 3M Rank",
    "isin": "ISIN",
}

NUM_COLS = {
    "close", "dma_20", "dma_50", "dma_100", "dma_200",
    "week_52_high", "market_cap_cr", "median_turnover_cr",
    "sharpe_6", "sharpe_3"
}
PCT_COLS = {"ROC_annual", "ROC_3", "ROC_6", "away_52wh"}
INT_COLS = {"sharpe_6_rank", "sharpe_3_rank", "Avg_sharpe_6_3_Rank", "total_circuit_hits_3m"}


def _write_excel_sheet(ws, df, title, header_fill, alt_fill, long_months=6, short_months=3):
    cols = list(df.columns)
    today_str = datetime.today().strftime("%d %b %Y")
    
    # Set headers map with dynamic months labels
    friendly_headers = dict(FRIENDLY_HEADERS)
    friendly_headers["sharpe_6"] = f"Sharpe {long_months}M"
    friendly_headers["sharpe_3"] = f"Sharpe {short_months}M"
    friendly_headers["ROC_6"] = f"{long_months}M ROC %"
    friendly_headers["ROC_3"] = f"{short_months}M ROC %"
    friendly_headers["sharpe_6_rank"] = f"Sharpe {long_months}M Rank"
    friendly_headers["sharpe_3_rank"] = f"Sharpe {short_months}M Rank"
    friendly_headers["Avg_sharpe_6_3_Rank"] = f"Avg Sharpe {long_months}M/{short_months}M Rank"

    # Merge title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(row=1, column=1, value=f"{title}  |  Screened: {today_str}")
    title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_cell.fill = header_fill
    title_cell.alignment = CENTER

    # Write column headers
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=ci, value=friendly_headers.get(col, col))
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)

    # Write data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=3):
        fill = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(cols, start=1):
            val = row[col]
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER
            cell.fill = fill

            if col in PCT_COLS:
                # Store percentage as fraction, Excel displays it as percentage
                cell.value = float(val) / 100.0 if (pd.notna(val) and val is not None) else None
                cell.number_format = "0.00%"
                cell.alignment = CENTER
                cell.font = DATA_FONT
            elif col in NUM_COLS:
                cell.value = float(val) if (pd.notna(val) and val is not None) else None
                cell.number_format = "#,##0.00"
                cell.alignment = CENTER
                cell.font = DATA_FONT
            elif col in INT_COLS:
                cell.value = int(val) if (pd.notna(val) and val is not None) else None
                cell.number_format = "0"
                cell.alignment = CENTER
                cell.font = DATA_FONT
            else:
                cell.value = str(val) if (pd.notna(val) and val is not None) else ""
                cell.alignment = LEFT
                cell.font = DATA_FONT

    # Freeze panes and enable filters
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20


def export_screener_to_excel(
    df_all: pd.DataFrame,
    df_filtered: pd.DataFrame,
    target_date: date,
    long_months: int = 6,
    short_months: int = 3
) -> str:
    """
    Exports screener results to reports/<date>.xlsx.
    Returns:
        Absolute file path to the saved Excel sheet.
    """
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    
    file_name = f"{target_date.strftime('%Y-%m-%d')}.xlsx"
    out_path = reports_dir / file_name
    
    wb = Workbook()
    
    # Sheet 1: All Passed Stocks
    ws1 = wb.active
    ws1.title = "All Stocks"
    _write_excel_sheet(
        ws1,
        df_all,
        title=f"NSE Sharpe Screener -- All Ranked Stocks ({long_months}M / {short_months}M)",
        header_fill=HEADER_FILL,
        alt_fill=ALT_FILL,
        long_months=long_months,
        short_months=short_months
    )
    
    # Sheet 2: High Conviction Filtered Watchlist
    ws2 = wb.create_sheet(title="Filtered")
    if df_filtered.empty:
        ws2.cell(row=1, column=1, value="No stocks matched the strict filter criteria.")
    else:
        _write_excel_sheet(
            ws2,
            df_filtered,
            title=f"Filtered Watchlist  |  3M ROC > 20%  |  Close >= 75% of 52WH  |  Circuit Hits <= 10",
            header_fill=FILTER_FILL,
            alt_fill=ALT_FILL_FILTER,
            long_months=long_months,
            short_months=short_months
        )
        
    wb.save(str(out_path))
    logger.info(f"Screener reports successfully written to: {out_path.absolute()}")
    return str(out_path.absolute())
